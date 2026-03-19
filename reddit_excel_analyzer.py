"""
reddit_excel_analyzer.py

Scrapes top hot posts from r/excel via Reddit's public JSON API
and saves results to a formatted Excel workbook.
"""

import time
import datetime
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

REDDIT_URL = "https://www.reddit.com/r/excel/hot.json"
USER_AGENT = "reddit-excel-analyzer/1.0 (educational script)"
MAX_POSTS = 300
OUTPUT_FILE = "reddit_excel_analysis.xlsx"

CATEGORY_KEYWORDS = {
    "formulas": [
        "vlookup", "xlookup", "sumif", "sumifs", "countif", "countifs",
        "lambda", "index", "match", "array formula", "dynamic array",
        "filter", "sort", "unique", "sequence", "iferror", "let", "offset",
        "formulatext", "indirect", " if(", " if ", "nested if",
    ],
    "vba": [
        "vba", "macro", " sub ", "sub(", "userform", "module",
        "automation", "xlsm", "personal.xlsb", "activex", "msgbox",
    ],
    "powerquery": [
        "power query", "power bi", "m code", "dax", "get & transform",
        "query editor", "unpivot", "powerquery", "powerbi",
    ],
    "charts": [
        "chart", "graph", "plot", "conditional format", "heatmap",
        "visualization", "sparkline", "axis", "legend", "gauge",
    ],
    "pivot": [
        "pivot", "slicer", "getpivotdata", "calculated field",
        "pivot table", "pivottable",
    ],
    "data": [
        "duplicate", "dedupe", "clean", "trim", "text to column",
        "flash fill", "substitute", " left(", " right(", " mid(",
        " len(", "datevalue", "isblank", "number format", "date format",
        "text function", "concatenate", "concat",
    ],
    "perf": [
        "slow", "performance", "large file", "memory", "crash", "freeze",
        "volatile", "optimize", "xlsb", "file size", "speed", "lag",
    ],
    "sharing": [
        "share", "protect", "password", "lock", "sharepoint", "onedrive",
        "co-author", "coauthor", "permission", "read only", "teams",
        "collaborate", "shared workbook",
    ],
}

CATEGORY_COLORS = {
    "formulas":   "DDEEFF",
    "vba":        "FFE4B5",
    "powerquery": "D5E8D4",
    "charts":     "FFF2CC",
    "pivot":      "E1D5E7",
    "data":       "DAE8FC",
    "perf":       "F8CECC",
    "sharing":    "D6EAF8",
    "other":      "F5F5F5",
}

HEADER_FILL = "404040"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_COLOR = "F0F0F0"

# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def fetch_posts(max_posts: int = MAX_POSTS) -> list:
    """Fetch up to max_posts hot posts from r/excel with pagination."""
    posts = []
    after = None
    headers = {"User-Agent": USER_AGENT}

    while len(posts) < max_posts:
        params = {"limit": 100}
        if after:
            params["after"] = after

        data = _get_with_retry(REDDIT_URL, headers=headers, params=params)
        if data is None:
            print("Warning: Could not fetch more posts; using what was collected.")
            break

        children = data.get("data", {}).get("children", [])
        if not children:
            break

        posts.extend(children)
        after = data.get("data", {}).get("after")
        if not after:
            break

        time.sleep(1)  # be polite to the API

    return posts[:max_posts]


def _get_with_retry(url: str, headers: dict, params: dict, retries: int = 3):
    """GET request with up to `retries` retries on failure."""
    delay = 2
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=15)
            resp.raise_for_status()
            return resp.json()
        except Exception as exc:
            if attempt < retries - 1:
                print(f"Request failed ({exc}), retrying in {delay}s...")
                time.sleep(delay)
                delay *= 2
            else:
                print(f"Request failed after {retries} attempts: {exc}")
                return None

# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify_post(title: str) -> str:
    """Return a category string based on keyword matching in the title."""
    lower = title.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in lower:
                return category
    return "other"

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_post(raw: dict, rank: int) -> dict | None:
    """Extract fields from a raw Reddit post dict; return None on error."""
    try:
        d = raw["data"]
        created = datetime.datetime.utcfromtimestamp(d["created_utc"])
        date_str = created.strftime("%Y-%m-%d")
        url = "https://www.reddit.com" + d["permalink"]
        title = d.get("title", "")
        return {
            "rank": rank,
            "title": title,
            "category": classify_post(title),
            "score": d.get("score", 0),
            "comments": d.get("num_comments", 0),
            "flair": d.get("link_flair_text") or "",
            "url": url,
            "date": date_str,
        }
    except Exception as exc:
        print(f"  Warning: skipped post at rank {rank} ({exc})")
        return None

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _header_style():
    return {
        "font": Font(bold=True, color=HEADER_FONT_COLOR, size=12),
        "fill": PatternFill("solid", fgColor=HEADER_FILL),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _apply_header(ws, headers: list, col_widths: list):
    style = _header_style()
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20


def _category_fill(category: str) -> PatternFill:
    color = CATEGORY_COLORS.get(category, CATEGORY_COLORS["other"])
    return PatternFill("solid", fgColor=color)


def _write_posts_sheet(ws, posts: list):
    """Write data rows to an 'All Posts' style sheet."""
    headers = ["Rank", "Title", "Category", "Score", "Comments", "Flair", "Post URL", "Date Posted"]
    col_widths = [6, 60, 14, 8, 10, 20, 50, 14]
    _apply_header(ws, headers, col_widths)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    for row_idx, post in enumerate(posts, start=2):
        bg_color = ALT_ROW_COLOR if row_idx % 2 == 0 else "FFFFFF"
        default_fill = PatternFill("solid", fgColor=bg_color)
        cat_fill = _category_fill(post["category"])

        values = [
            post["rank"], post["title"], post["category"],
            post["score"], post["comments"], post["flair"],
            post["url"], post["date"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 2))
            # Category column gets unique color; others get alternating
            if col_idx == 3:
                cell.fill = cat_fill
            else:
                cell.fill = default_fill

        ws.row_dimensions[row_idx].height = 30

# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(posts: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ---- Sheet 1: All Posts ----
    ws_all = wb.create_sheet("All Posts")
    _write_posts_sheet(ws_all, posts)

    # ---- Sheet 2: Category Summary ----
    ws_cat = wb.create_sheet("Category Summary")
    _build_summary_sheet(ws_cat, posts)

    # ---- Sheet 3: Top 20 Posts ----
    ws_top = wb.create_sheet("Top 20 Posts")
    top20 = sorted(posts, key=lambda p: p["score"], reverse=True)[:20]
    _write_posts_sheet(ws_top, top20)

    # ---- Sheet 4: Charts ----
    ws_charts = wb.create_sheet("Charts")
    _build_charts_sheet(ws_charts, posts)

    return wb


def _build_summary_sheet(ws, posts: list):
    headers = ["Category", "Post Count", "% of Total", "Avg Score", "Avg Comments",
               "Top Post Title", "Top Post Score"]
    col_widths = [14, 12, 12, 12, 14, 60, 14]
    _apply_header(ws, headers, col_widths)

    total = len(posts)
    by_cat = defaultdict(list)
    for p in posts:
        by_cat[p["category"]].append(p)

    rows = []
    for cat, cat_posts in by_cat.items():
        top = max(cat_posts, key=lambda p: p["score"])
        rows.append({
            "category": cat,
            "count": len(cat_posts),
            "pct": len(cat_posts) / total * 100 if total else 0,
            "avg_score": sum(p["score"] for p in cat_posts) / len(cat_posts),
            "avg_comments": sum(p["comments"] for p in cat_posts) / len(cat_posts),
            "top_title": top["title"],
            "top_score": top["score"],
        })

    rows.sort(key=lambda r: r["count"], reverse=True)

    for row_idx, r in enumerate(rows, start=2):
        bg_color = ALT_ROW_COLOR if row_idx % 2 == 0 else "FFFFFF"
        default_fill = PatternFill("solid", fgColor=bg_color)
        cat_fill = _category_fill(r["category"])

        values = [
            r["category"], r["count"], f"{r['pct']:.1f}%",
            f"{r['avg_score']:.1f}", f"{r['avg_comments']:.1f}",
            r["top_title"], r["top_score"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 6))
            cell.fill = cat_fill if col_idx == 1 else default_fill

        ws.row_dimensions[row_idx].height = 30


def _build_charts_sheet(ws, posts: list):
    """Write a data table and two bar charts to the Charts sheet."""
    # Build summary data table on this sheet for chart references
    by_cat = defaultdict(list)
    for p in posts:
        by_cat[p["category"]].append(p)

    rows = sorted(by_cat.items(), key=lambda x: len(x[1]), reverse=True)

    # Header row
    ws["A1"] = "Category"
    ws["B1"] = "Post Count"
    ws["C1"] = "Avg Score"
    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 16

    for i, (cat, cat_posts) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=len(cat_posts))
        avg_score = sum(p["score"] for p in cat_posts) / len(cat_posts)
        ws.cell(row=i, column=3, value=round(avg_score, 1))

    num_cats = len(rows)
    data_rows = num_cats + 1  # includes header

    # Chart 1: Post Count by Category
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Post Count by Category"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Category"
    chart1.width = 20
    chart1.height = 12

    data1 = Reference(ws, min_col=2, min_row=1, max_row=data_rows)
    cats1 = Reference(ws, min_col=1, min_row=2, max_row=data_rows)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "E2")

    # Chart 2: Avg Score by Category
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Avg Score by Category"
    chart2.y_axis.title = "Avg Score"
    chart2.x_axis.title = "Category"
    chart2.width = 20
    chart2.height = 12

    data2 = Reference(ws, min_col=3, min_row=1, max_row=data_rows)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1)
    ws.add_chart(chart2, "E22")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Fetching posts from r/excel...")
    raw_posts = fetch_posts()
    print(f"Fetched {len(raw_posts)} raw posts.")

    posts = []
    for rank, raw in enumerate(raw_posts, start=1):
        parsed = parse_post(raw, rank)
        if parsed:
            posts.append(parsed)

    print(f"Parsed {len(posts)} posts successfully.\n")

    # Console summary
    from collections import Counter
    counts = Counter(p["category"] for p in posts)
    print(f"{'Category':<14} {'Count':>6}")
    print("-" * 22)
    for cat, count in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"{cat:<14} {count:>6}")
    print()

    print("Building Excel workbook...")
    wb = build_workbook(posts)
    wb.save(OUTPUT_FILE)
    print(f"Done! {len(posts)} posts saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
